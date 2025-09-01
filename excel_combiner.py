"""
CombineXL

Description:
    A GUI-based utility for combining multiple Excel files (.xlsx) into a single,
    master workbook. The tool is designed for performance and usability,
    featuring a responsive, non-blocking interface that preserves all cell
    formatting, styles, and merged regions.

The program executes the following numbered steps:
  1. Launches a GUI and prompts the user to select multiple input Excel files.
  2. Presents a custom dialog where the user can reorder the selected files
     to define the exact sequence of combination.
  3. Asks the user to choose a performance optimization method: a Standard,
     cell-by-cell copy, or a significantly faster Optimized method that uses
     style caching.
  4. Prompts for a save location, intelligently defaulting to the input
     directory with a descriptive, timestamped filename.
  5. Asks for all other combination parameters via dialogs:
     - Number of header rows to retain from the first file.
     - Number of header rows to discard from all subsequent files.
     - Whether to include the source filename in a new first column.
     - Whether to preserve formulas or copy only their static, calculated values.
  6. Launches the main file combination logic in a separate, non-blocking
     worker thread to keep the user interface responsive.
  7. Displays a real-time progress window showing the current status and
     overall progress, complete with a "Cancel" button.
  8. The worker thread iterates through each file, meticulously copying data
     and all associated formatting using the user-selected copy method.
  9. Before processing each file, the thread checks for a cancellation signal
     to allow for a graceful exit.
 10. Upon completion, error, or cancellation, a final summary report is
     displayed in a GUI dialog and printed to the console.

Usage:
    - Ensure required libraries are installed:
          pip install openpyxl
    - Run the script from a terminal or by executing the file directly:
          python xls_combiner.py

Author:     Vitalii Starosta
GitHub:     https://github.com/sztaroszta
License:    GNU Affero General Public License v3 (AGPLv3)
"""

import os
import sys
import tkinter as tk
from tkinter import filedialog, simpledialog, messagebox, ttk
from copy import copy
from datetime import datetime
import openpyxl
from openpyxl.utils import get_column_letter
import threading
import queue

# --- Dialog Classes ---

class FileOrderDialog(tk.Toplevel):
    """
    A custom Toplevel dialog window that allows the user to reorder a list of files.
    
    The dialog displays a list of file basenames and provides "Move Up" and
    "Move Down" buttons for reordering. The final, ordered list of full file
    paths is returned upon closing or clicking "OK".
    """
    def __init__(self, master, files):
        """
        Initializes the FileOrderDialog.

        Args:
            master (tk.Tk or tk.Toplevel): The parent window for this dialog.
            files (list[str]): A list of full file paths to be ordered.
        """
        super().__init__(master)
        self.title("Order Files for Combination")
        self.files = list(files)
        self.result = None

        max_length = max(len(os.path.basename(file)) for file in self.files) if self.files else 20
        width = max_length + 4

        list_frame = tk.Frame(self)
        list_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=10, pady=10)
        
        self.listbox = tk.Listbox(list_frame, selectmode=tk.SINGLE, width=width, height=15)
        for file in self.files:
            self.listbox.insert(tk.END, os.path.basename(file))
        self.listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        scrollbar = tk.Scrollbar(list_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        self.listbox.config(yscrollcommand=scrollbar.set)
        scrollbar.config(command=self.listbox.yview)

        button_frame = tk.Frame(self)
        button_frame.pack(side=tk.RIGHT, fill=tk.Y, padx=10, pady=10)
        
        tk.Button(button_frame, text="Move Up", command=self.move_up).pack(fill=tk.X, pady=2)
        tk.Button(button_frame, text="Move Down", command=self.move_down).pack(fill=tk.X, pady=2)
        tk.Button(button_frame, text="OK", command=self.on_ok, font=('Helvetica', 10, 'bold')).pack(fill=tk.X, pady=(10, 2))

        if self.listbox.size() > 0:
            self.listbox.selection_set(0)

        self.protocol("WM_DELETE_WINDOW", self.on_close)
        self.transient(master)
        self.grab_set()
        self.wait_window(self)

    def move_up(self):
        """Moves the selected file one position up in the list."""
        selection = self.listbox.curselection()
        if not selection or selection[0] == 0: return
        idx = selection[0]
        self.files[idx], self.files[idx - 1] = self.files[idx - 1], self.files[idx]
        self.update_listbox()
        self.listbox.selection_set(idx - 1)

    def move_down(self):
        """Moves the selected file one position down in the list."""
        selection = self.listbox.curselection()
        if not selection or selection[0] == self.listbox.size() - 1: return
        idx = selection[0]
        self.files[idx], self.files[idx + 1] = self.files[idx + 1], self.files[idx]
        self.update_listbox()
        self.listbox.selection_set(idx + 1)

    def update_listbox(self):
        """Refreshes the listbox display based on the current order of `self.files`."""
        self.listbox.delete(0, tk.END)
        for file in self.files:
            self.listbox.insert(tk.END, os.path.basename(file))

    def on_ok(self):
        """Sets the result to the final ordered list and closes the dialog."""
        self.result = self.files
        self.destroy()

    def on_close(self):
        """Sets the result to None to indicate cancellation and closes the dialog."""
        self.result = None
        self.destroy()

class ProgressManager:
    """
    Manages the Toplevel progress window UI.
    
    This class displays the current status text, a progress bar, and a
    cancel button. It also mirrors progress updates to the system console.
    """
    def __init__(self, parent, title, total_steps, cancel_event):
        """
        Initializes the ProgressManager window.

        Args:
            parent (tk.Tk or tk.Toplevel): The parent window.
            title (str): The title for the progress window.
            total_steps (int): The total number of steps for the progress bar.
            cancel_event (threading.Event): The event to set when cancellation is requested.
        """
        self.parent = parent
        self.total_steps = total_steps
        self.cancel_event = cancel_event
        self.window = tk.Toplevel(parent)
        self.window.title(title)
        self.window.resizable(False, False)
        self.window.protocol("WM_DELETE_WINDOW", self.request_cancel)
        self.status_label = tk.Label(self.window, text="Initializing...", padx=20, pady=10, width=60)
        self.status_label.pack()
        self.progress_bar = ttk.Progressbar(self.window, orient="horizontal", length=400, mode="determinate", maximum=total_steps)
        self.progress_bar.pack(padx=20, pady=5)
        self.cancel_button = tk.Button(self.window, text="Cancel", command=self.request_cancel, width=10)
        self.cancel_button.pack(pady=10)
        self.parent.update_idletasks()

    def update(self, current_step, status_text):
        """
        Updates the GUI and terminal progress indicators.

        Args:
            current_step (int): The current step in the process.
            status_text (str): The text to display as the current status.
        """
        self.progress_bar['value'] = current_step
        self.status_label.config(text=status_text)
        progress_percent = (current_step / self.total_steps) * 100
        bar_length = 30
        filled_length = int(bar_length * current_step // self.total_steps)
        bar = '█' * filled_length + '-' * (bar_length - filled_length)
        terminal_text = f"\rProgress: |{bar}| {progress_percent:.1f}% ({current_step}/{self.total_steps}) - Processing..."
        sys.stdout.write(terminal_text)
        sys.stdout.flush()
        self.parent.update_idletasks()

    def request_cancel(self):
        """Flags the operation for cancellation by setting the threading event."""
        if messagebox.askyesno("Confirm Cancel", "Are you sure you want to cancel the operation?"):
            self.status_label.config(text="Cancellation requested...")
            self.cancel_event.set()

    def close(self):
        """Closes the progress window and prints a final newline to the terminal."""
        sys.stdout.write('\n')
        self.window.destroy()

# --- Copying Functions ---

def copy_cell_v1(src_cell, tgt_cell):
    """
    Standard (slow) method: Copies cell value and all style attributes individually.
    
    This creates new style objects for every cell, ensuring compatibility but
    incurring significant performance overhead.

    Args:
        src_cell (openpyxl.cell.Cell): The source cell to copy from.
        tgt_cell (openpyxl.cell.Cell): The target cell to copy to.
    """
    tgt_cell.value = src_cell.value
    if src_cell.has_style:
        tgt_cell.font = copy(src_cell.font)
        tgt_cell.border = copy(src_cell.border)
        tgt_cell.fill = copy(src_cell.fill)
        tgt_cell.number_format = src_cell.number_format
        tgt_cell.protection = copy(src_cell.protection)
        tgt_cell.alignment = copy(src_cell.alignment)
    if src_cell.hyperlink:
        tgt_cell.hyperlink = copy(src_cell.hyperlink)
    if src_cell.comment:
        tgt_cell.comment = copy(src_cell.comment)

def copy_row_v1(src_sheet, tgt_sheet, src_row_idx, tgt_row_idx, include_filename=False, filename=None):
    """Standard (slow) method: Copies an entire row by iterating through each cell."""
    if src_row_idx in src_sheet.row_dimensions:
        tgt_sheet.row_dimensions[tgt_row_idx].height = src_sheet.row_dimensions[src_row_idx].height
    col_offset = 1 if include_filename else 0
    if include_filename:
        tgt_sheet.cell(row=tgt_row_idx, column=1).value = filename
    for col_idx in range(1, src_sheet.max_column + 1):
        src_cell = src_sheet.cell(row=src_row_idx, column=col_idx)
        tgt_cell = tgt_sheet.cell(row=tgt_row_idx, column=col_idx + col_offset)
        copy_cell(src_cell, tgt_cell)

style_cache = {}

def copy_cell_v2(src_cell, tgt_cell):
    """
    Optimized (fast) method: Copies cell value and reuses style objects via a cache.
    
    This method performs a full style copy only once per unique style, and all
    subsequent cells with the same style get a fast assignment from the cache.

    Args:
        src_cell (openpyxl.cell.Cell): The source cell to copy from.
        tgt_cell (openpyxl.cell.Cell): The target cell to copy to.
    """
    tgt_cell.value = src_cell.value
    if src_cell.has_style:
        style_key = src_cell._style
        if style_key in style_cache:
            tgt_cell._style = style_cache[style_key]
        else:
            tgt_cell.font = copy(src_cell.font)
            tgt_cell.border = copy(src_cell.border)
            tgt_cell.fill = copy(src_cell.fill)
            tgt_cell.number_format = src_cell.number_format
            tgt_cell.protection = copy(src_cell.protection)
            tgt_cell.alignment = copy(src_cell.alignment)
            style_cache[style_key] = tgt_cell._style
    if src_cell.hyperlink:
        tgt_cell.hyperlink = copy(src_cell.hyperlink)
    if src_cell.comment:
        tgt_cell.comment = copy(src_cell.comment)

def copy_row_v2(src_sheet, tgt_sheet, src_row_idx, tgt_row_idx, include_filename=False, filename=None):
    """Optimized (fast) method: Copies an entire row by iterating through each cell."""
    if src_row_idx in src_sheet.row_dimensions:
        tgt_sheet.row_dimensions[tgt_row_idx].height = src_sheet.row_dimensions[src_row_idx].height
    col_offset = 1 if include_filename else 0
    if include_filename:
        tgt_sheet.cell(row=tgt_row_idx, column=1).value = filename
    for col_idx in range(1, src_sheet.max_column + 1):
        src_cell = src_sheet.cell(row=src_row_idx, column=col_idx)
        tgt_cell = tgt_sheet.cell(row=tgt_row_idx, column=col_idx + col_offset)
        copy_cell(src_cell, tgt_cell)

copy_cell = None
copy_row = None

def copy_merged_cells(src_sheet, tgt_sheet, src_start_row, src_end_row, tgt_start_row, include_filename=False):
    """
    Copies merged cell ranges from a source worksheet to a target worksheet.

    Args:
        src_sheet (openpyxl.worksheet.Worksheet): The source worksheet.
        tgt_sheet (openpyxl.worksheet.Worksheet): The target worksheet.
        src_start_row (int): The starting row of the source range to check for merges.
        src_end_row (int): The ending row of the source range.
        tgt_start_row (int): The row in the target sheet where the new range begins.
        include_filename (bool): If True, shifts column indices by 1.
    """
    col_offset = 1 if include_filename else 0
    for merge_range in src_sheet.merged_cells.ranges:
        if merge_range.min_row >= src_start_row and merge_range.max_row <= src_end_row:
            new_min_row = (merge_range.min_row - src_start_row) + tgt_start_row
            new_max_row = (merge_range.max_row - src_start_row) + tgt_start_row
            new_min_col = merge_range.min_col + col_offset
            new_max_col = merge_range.max_col + col_offset
            try:
                tgt_sheet.merge_cells(start_row=new_min_row, start_column=new_min_col, end_row=new_max_row, end_column=new_max_col)
            except Exception as e:
                print(f"Warning: Could not merge range {merge_range}: {e}")

# --- Worker Function ---
def combine_excel_files_worker(files, output_file, heading_rows, delete_rows, include_filename, preserve_formulas, progress_queue, cancel_event):
    """
    Performs the Excel combination in a worker thread.
    
    This function contains the core logic for iterating through files, copying
    data, and saving the final workbook. It communicates progress and results
    back to the main thread via a queue.

    Args:
        files (list[str]): The ordered list of Excel files to combine.
        output_file (str): The path to save the combined Excel file.
        heading_rows (int): The number of header rows from the first file to keep.
        delete_rows (int): The number of rows to discard from subsequent files.
        include_filename (bool): If True, adds a filename column.
        preserve_formulas (bool): If True, preserves formulas; otherwise, copies values.
        progress_queue (queue.Queue): Queue for sending status updates.
        cancel_event (threading.Event): Event to check for cancellation requests.
    """
    try:
        combined_wb = openpyxl.Workbook()
        combined_ws = combined_wb.active
        combined_ws.title = "Combined Data"
        tgt_row_idx = 1
        
        load_data_only = not preserve_formulas
        style_cache.clear()

        for i, file_path in enumerate(files):
            if cancel_event.is_set():
                result = {'status': 'cancelled', 'message': 'Operation cancelled by user.'}
                progress_queue.put({'type': 'result', 'data': result})
                return

            status_text = f"Processing file {i + 1}/{len(files)}: {os.path.basename(file_path)}"
            progress_queue.put({'type': 'progress', 'step': i, 'status': status_text})

            wb = openpyxl.load_workbook(file_path, data_only=load_data_only)
            ws = wb.active
            filename = os.path.basename(file_path)

            if i == 0:
                for src_row in range(1, heading_rows + 1):
                    copy_row(ws, combined_ws, src_row, tgt_row_idx, include_filename, filename)
                    tgt_row_idx += 1
                copy_merged_cells(ws, combined_ws, 1, heading_rows, 1, include_filename)
                
                data_start_row = heading_rows + 1
                for src_row in range(data_start_row, ws.max_row + 1):
                    copy_row(ws, combined_ws, src_row, tgt_row_idx, include_filename, filename)
                    tgt_row_idx += 1
                copy_merged_cells(ws, combined_ws, data_start_row, ws.max_row, heading_rows + 1, include_filename)
                
                col_offset = 1 if include_filename else 0
                for col_letter, dim in ws.column_dimensions.items():
                    tgt_col_letter = get_column_letter(openpyxl.utils.column_index_from_string(col_letter) + col_offset)
                    if dim.width:
                        combined_ws.column_dimensions[tgt_col_letter].width = dim.width
            else:
                start_row = delete_rows + 1
                if start_row > ws.max_row: continue

                block_tgt_start_row = tgt_row_idx
                for src_row in range(start_row, ws.max_row + 1):
                    copy_row(ws, combined_ws, src_row, tgt_row_idx, include_filename, filename)
                    tgt_row_idx += 1
                copy_merged_cells(ws, combined_ws, start_row, ws.max_row, block_tgt_start_row, include_filename)
        
        progress_queue.put({'type': 'progress', 'step': len(files), 'status': 'Saving combined file...'})
        combined_wb.save(output_file)

        result = {'status': 'success', 'message': f'Successfully combined {len(files)} files.'}
        progress_queue.put({'type': 'result', 'data': result})

    except Exception as e:
        result = {'status': 'error', 'message': f"An error occurred: {e}"}
        progress_queue.put({'type': 'result', 'data': result})

# --- Main Application Class ---
class App:
    """
    The main application class that orchestrates the GUI, user input, and the
    background worker thread.
    """
    def __init__(self, root):
        """
        Initializes and runs the main application.

        Args:
            root (tk.Tk): The root Tkinter window, which will be managed by the app.
        """
        self.root = root
        self.progress_manager = None
        self.run()

    def get_user_input(self):
        """
        Handles all initial user dialogs to get processing parameters.

        This method guides the user through selecting files, ordering them,
        choosing options, and selecting an output path.

        Returns:
            bool: True if the user completed all prompts, False if they cancelled.
        """
        file_list = filedialog.askopenfilenames(title="Select Input Excel Files (.xlsx)", filetypes=[("Excel files", "*.xlsx")])
        if not file_list: return False
        
        self.root.deiconify()
        order_dialog = FileOrderDialog(self.root, file_list)
        self.ordered_files = order_dialog.result
        self.root.withdraw()

        if not self.ordered_files: return False

        use_version2 = messagebox.askyesno(
            title="Select Cell Copying Method",
            message="Please choose the method for copying cell styles.\n\n"
                    "• Optimized (Recommended): Uses style caching for a significant speed boost.\n\n"
                    "• Standard: Slower, direct-copy method for baseline compatibility.\n\n"
                    "Do you want to use the Optimized method?"
        )

        global copy_cell, copy_row
        if use_version2:
            copy_cell = copy_cell_v2
            copy_row = copy_row_v2
            print("Using optimized copy functions (Style Caching).")
        else:
            copy_cell = copy_cell_v1
            copy_row = copy_row_v1
            print("Using standard copy functions.")
        
        first_file_path = self.ordered_files[0]
        output_dir = os.path.dirname(first_file_path)
        first_file_stem = os.path.splitext(os.path.basename(first_file_path))[0]
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        default_filename = f"{first_file_stem}_combined_{timestamp}.xlsx"

        self.output_file = filedialog.asksaveasfilename(
            title="Select Output Excel File (.xlsx)",
            initialdir=output_dir,
            initialfile=default_filename,
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")]
        )
        if not self.output_file: return False

        self.heading_rows = simpledialog.askinteger("Header Rows (First File)", "Enter header rows to retain from the first file:", initialvalue=1, minvalue=1)
        if self.heading_rows is None: return False

        self.delete_rows = simpledialog.askinteger("Delete Rows (Other Files)", "Enter heading rows to delete from other files:", initialvalue=1, minvalue=0)
        if self.delete_rows is None: return False

        self.include_filename = messagebox.askyesno("Include Filename", "Add original filename as the first column?")
        
        self.preserve_formulas = messagebox.askyesno(
            title="Preserve Formulas?",
            message="Do you want to preserve formulas?\n\n"
                    "• 'Yes' will keep formulas (e.g., '=A1+B1'), but may cause #REF! errors.\n"
                    "• 'No' will copy only the calculated values (e.g., '123'), ensuring data is static."
        )
        return True

    def start_processing(self):
        """Prepares for and launches the background processing thread."""
        print("\n--- Settings ---")
        print(f"  Output file: {self.output_file}")
        print(f"  Files to combine: {len(self.ordered_files)}")
        print(f"  Header rows (first file): {self.heading_rows}")
        print(f"  Delete rows (other files): {self.delete_rows}")
        print(f"  Include filename column: {'Yes' if self.include_filename else 'No'}")
        print(f"  Preserve Formulas: {'Yes' if self.preserve_formulas else 'No'}")
        print("------------------\n")

        self.progress_queue = queue.Queue()
        self.cancel_event = threading.Event()
        
        self.progress_manager = ProgressManager(self.root, "Combining Files...", len(self.ordered_files), self.cancel_event)
        
        self.worker_thread = threading.Thread(
            target=combine_excel_files_worker,
            args=(self.ordered_files, self.output_file, self.heading_rows, self.delete_rows, self.include_filename, self.preserve_formulas, self.progress_queue, self.cancel_event)
        )
        self.worker_thread.start()
        self.root.after(100, self.check_queue)

    def check_queue(self):
        """
        Periodically checks the queue for messages from the worker thread.
        
        This method is the bridge between the background thread and the main
        GUI thread. It updates the progress bar and triggers the final report.
        """
        try:
            while True:
                message = self.progress_queue.get(block=False)
                if message['type'] == 'progress':
                    self.progress_manager.update(message['step'], message['status'])
                elif message['type'] == 'result':
                    self.on_task_finished(message['data'])
                    return
        except queue.Empty:
            pass

        if self.worker_thread.is_alive():
            self.root.after(100, self.check_queue)
        else:
            self.on_task_finished({'status': 'error', 'message': 'The worker thread terminated unexpectedly.'})

    def on_task_finished(self, result):
        """
        Handles the final result from the worker and displays a summary.
        
        Args:
            result (dict): A dictionary with 'status' and 'message' keys from the worker.
        """
        if self.progress_manager:
            self.progress_manager.close()
        final_message = result.get('message', 'No message provided.')
        print("\n--- Operation Summary ---")
        print(f"Status: {result['status'].title()}")
        print(f"Message: {final_message}")
        print("-------------------------\n")
        status_map = {
            'error': messagebox.showerror,
            'cancelled': messagebox.showwarning,
            'success': messagebox.showinfo
        }
        title = result['status'].title()
        if result['status'] == 'success':
            message_body = f"{final_message}\n\nCombined file saved to:\n{self.output_file}"
        else:
            message_body = final_message
        status_map.get(result['status'], messagebox.showinfo)(title, message_body)
        self.root.destroy()

    def run(self):
        """Main execution flow of the application."""
        print("CombineXL")
        print("=" * 40)
        if self.get_user_input():
            self.start_processing()
        else:
            print("\nOperation cancelled during setup. Exiting...")
            self.root.destroy()

if __name__ == "__main__":
    try:
        root = tk.Tk()
        # Make the root window fully transparent and withdrawn. This allows it
        # to act as a parent for dialogs without ever being visible.
        root.attributes('-alpha', 0.0)
        root.withdraw()
        
        app = App(root)
        root.mainloop()
    except tk.TclError as e:
        print(f"Failed to start GUI application: {e}")
        print("This script requires a graphical desktop environment to run.")